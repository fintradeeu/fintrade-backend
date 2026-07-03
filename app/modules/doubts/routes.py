"""Doubts module — FastAPI routes."""

import io
from datetime import datetime, timezone
from typing import List

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import get_current_user, require_roles
from app.db.database import get_db
from app.modules.auth.models import User
from app.modules.doubts import services
from app.modules.doubts.schemas import (
    DoubtFormCreate,
    DoubtFormResponse,
    DoubtFormUpdate,
    DoubtSubmissionCreate,
    DoubtSubmissionResponse,
)

router = APIRouter(prefix="/doubts", tags=["Doubt Solving Panel"])


# ── Admin / Faculty / SuperAdmin endpoints ───────────────────────────────────

@router.post("/forms", response_model=DoubtFormResponse)
async def create_form(
    body: DoubtFormCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "faculty", "superadmin"])),
):
    """Create a new doubt form for a batch."""
    form = await services.create_doubt_form(db, body.model_dump(), creator_id=current_user.id)
    response = DoubtFormResponse.model_validate(form)
    response.submission_count = len(form.submissions)
    return response


@router.get("/forms", response_model=List[DoubtFormResponse])
async def list_forms(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "faculty", "superadmin"])),
):
    """List all doubt forms."""
    forms = await services.list_doubt_forms(db)
    result = []
    for form in forms:
        r = DoubtFormResponse.model_validate(form)
        r.submission_count = len(form.submissions)
        result.append(r)
    return result


@router.put("/forms/{form_id}", response_model=DoubtFormResponse)
async def update_form(
    form_id: int,
    body: DoubtFormUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "faculty", "superadmin"])),
):
    """Update a doubt form."""
    form = await services.update_doubt_form(db, form_id, body.model_dump(exclude_none=True))
    if not form:
        raise HTTPException(status_code=404, detail="Doubt form not found")
    response = DoubtFormResponse.model_validate(form)
    response.submission_count = len(form.submissions)
    return response


@router.delete("/forms/{form_id}")
async def delete_form(
    form_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "faculty", "superadmin"])),
):
    """Delete a doubt form."""
    deleted = await services.delete_doubt_form(db, form_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Doubt form not found")
    return {"message": "Doubt form deleted successfully"}


@router.get("/forms/{form_id}/submissions", response_model=List[DoubtSubmissionResponse])
async def get_submissions(
    form_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "faculty", "superadmin"])),
):
    """List all student doubts for a specific form."""
    form = await services.get_doubt_form(db, form_id)
    if not form:
        raise HTTPException(status_code=404, detail="Doubt form not found")
    submissions = await services.get_form_submissions(db, form_id)
    return submissions


@router.get("/forms/{form_id}/export")
async def export_submissions_excel(
    form_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_roles(["admin", "faculty", "superadmin"])),
):
    """Export all doubts for a form as an Excel (.xlsx) file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    form = await services.get_doubt_form(db, form_id)
    if not form:
        raise HTTPException(status_code=404, detail="Doubt form not found")

    submissions = await services.get_form_submissions(db, form_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doubts"

    # Header style
    header_fill = PatternFill(start_color="0B2A5B", end_color="0B2A5B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["#", "Student Name", "Email", "Phone", "Batch", "Topic", "Doubt / Question", "Submitted At"]
    col_widths = [5, 25, 30, 15, 25, 25, 50, 22]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, sub in enumerate(submissions, start=2):
        student = sub.student
        batch_name = form.batch.name if form.batch else str(form.batch_id)
        submitted_str = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub.submitted_at else ""

        row_data = [
            row_idx - 1,
            student.full_name if student else "",
            student.email if student else "",
            getattr(student, "phone", "") or "",
            batch_name,
            sub.topic or "",
            sub.doubt_text,
            submitted_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Alternating row color
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="F0F4FF", end_color="F0F4FF", fill_type="solid"
                )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_title = form.title.replace(" ", "_").replace("/", "-")[:40]
    filename = f"doubts_{safe_title}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Student endpoints ────────────────────────────────────────────────────────

@router.get("/student/forms", response_model=List[DoubtFormResponse])
async def get_student_forms(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get active doubt forms available for the logged-in student."""
    forms = await services.get_forms_for_student(db, current_user.id)
    result = []
    for form in forms:
        r = DoubtFormResponse.model_validate(form)
        r.submission_count = 0  # Hide count from students
        result.append(r)
    return result


@router.post("/student/forms/{form_id}/submit", response_model=DoubtSubmissionResponse)
async def submit_doubt(
    form_id: int,
    body: DoubtSubmissionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Submit a doubt for a specific form."""
    form = await services.get_doubt_form(db, form_id)
    if not form:
        raise HTTPException(status_code=404, detail="Doubt form not found")

    now = datetime.now(timezone.utc)
    if not form.is_active or form.end_date < now:
        raise HTTPException(status_code=400, detail="This doubt form has expired or is no longer active")

    already_submitted = await services.has_student_submitted(db, form_id, current_user.id)
    if already_submitted:
        raise HTTPException(status_code=400, detail="You have already submitted a doubt for this form")

    submission = await services.submit_doubt(db, form_id, current_user.id, body.model_dump())
    return submission
