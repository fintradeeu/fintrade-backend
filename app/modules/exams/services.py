"""Exams module — service layer with 30-day reattempt restriction."""

from datetime import datetime, timedelta, timezone
from typing import List, Optional

from fastapi import HTTPException, status
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.modules.exams.models import (
    EntranceExam,
    ExamAnswer,
    ExamAttempt,
    ExamOption,
    ExamQuestion,
    ExamResult,
)
from app.utils.logger import get_logger

logger = get_logger(__name__)

REATTEMPT_DAYS = 30


# ── Helpers ──────────────────────────────────────────────────────────
async def _check_reattempt_allowed(db: AsyncSession, user_id: int, exam_id: int) -> None:
    """Raise 403 if the student failed and must wait 30 days."""
    # Find the most recent completed attempt
    result = await db.execute(
        select(ExamAttempt)
        .where(
            ExamAttempt.user_id == user_id,
            ExamAttempt.exam_id == exam_id,
            ExamAttempt.is_submitted == True,  # noqa: E712
        )
        .order_by(ExamAttempt.submitted_at.desc())
        .limit(1)
    )
    last_attempt = result.scalar_one_or_none()
    if last_attempt is None:
        return  # first attempt — allowed

    # Check if the last attempt passed
    result_row = await db.execute(
        select(ExamResult).where(ExamResult.attempt_id == last_attempt.id)
    )
    exam_result = result_row.scalar_one_or_none()

    if exam_result and exam_result.passed:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="You have already passed this exam.",
        )

    # Failed — enforce 30-day wait
    if last_attempt.submitted_at:
        next_allowed = last_attempt.submitted_at + timedelta(days=REATTEMPT_DAYS)
        if datetime.now(timezone.utc) < next_allowed:
            days_left = (next_allowed - datetime.now(timezone.utc)).days
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"You must wait {days_left} more day(s) before reattempting. Next attempt allowed after {next_allowed.strftime('%Y-%m-%d')}.",
            )


# ── Admin: create exam ──────────────────────────────────────────────
async def create_exam(db: AsyncSession, data: dict) -> EntranceExam:
    """Create an entrance exam with questions and options."""
    exam = EntranceExam(
        title=data["title"],
        description=data.get("description"),
        course_id=data["course_id"],
        duration_minutes=data.get("duration_minutes", 60),
        passing_score=data.get("passing_score", 60.0),
        is_active=data.get("is_active", True),
        start_time=data.get("start_time"),
        end_time=data.get("end_time"),
    )
    db.add(exam)
    await db.flush()

    for q_data in data.get("questions", []):
        question = ExamQuestion(
            exam_id=exam.id,
            question_text=q_data["question_text"],
            question_type=q_data.get("question_type", "mcq"),
            marks=q_data.get("marks", 1.0),
            negative_marks=q_data.get("negative_marks", 0.0),
            category=q_data.get("category"),
            order=q_data.get("order", 0),
            explanation=q_data.get("explanation"),
        )
        db.add(question)
        await db.flush()

        for opt_data in q_data.get("options", []):
            option = ExamOption(
                question_id=question.id,
                option_text=opt_data["option_text"],
                is_correct=opt_data.get("is_correct", False),
                order=opt_data.get("order", 0),
            )
            db.add(option)

    await db.flush()
    await db.refresh(exam)
    logger.info("exam_created", exam_id=exam.id, title=exam.title)
    return exam

async def create_course_exam(db: AsyncSession, data: dict) -> "CourseExam":
    from app.modules.exams.models import CourseExam, CourseExamQuestion, CourseExamOption
    exam = CourseExam(
        title=data["title"],
        description=data.get("description"),
        course_id=data["course_id"],
        module_id=data.get("module_id"),
        exam_type=data.get("exam_type", "course_final"),
        duration_minutes=data.get("duration_minutes", 60),
        passing_score=data.get("passing_score", 60.0),
        max_attempts=data.get("max_attempts", 3),
        is_active=data.get("is_active", True),
        start_time=data.get("start_time"),
        end_time=data.get("end_time"),
    )
    db.add(exam)
    await db.flush()

    for q_data in data.get("questions", []):
        question = CourseExamQuestion(
            exam_id=exam.id,
            question_text=q_data["question_text"],
            question_type=q_data.get("question_type", "mcq"),
            marks=q_data.get("marks", 1.0),
            negative_marks=q_data.get("negative_marks", 0.0),
            category=q_data.get("category"),
            order=q_data.get("order", 0),
            explanation=q_data.get("explanation"),
        )
        db.add(question)
        await db.flush()

        for opt_data in q_data.get("options", []):
            option = CourseExamOption(
                question_id=question.id,
                option_text=opt_data["option_text"],
                is_correct=opt_data.get("is_correct", False),
                order=opt_data.get("order", 0),
            )
            db.add(option)

    await db.flush()
    await db.refresh(exam)
    logger.info("course_exam_created", exam_id=exam.id, title=exam.title)
    return exam

async def update_exam(db: AsyncSession, exam_id: int, data: dict, is_course_exam: bool = False):
    from app.modules.exams.models import EntranceExam, CourseExam
    model = CourseExam if is_course_exam else EntranceExam
    exam = await db.get(model, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    
    if data.get("title") is not None:
        exam.title = data["title"]
    if data.get("description") is not None:
        exam.description = data["description"]
    if data.get("duration_minutes") is not None:
        exam.duration_minutes = data["duration_minutes"]
    if data.get("passing_score") is not None:
        exam.passing_score = data["passing_score"]
    if data.get("is_active") is not None:
        exam.is_active = data["is_active"]
        
    await db.flush()
    await db.refresh(exam)
    return exam


async def add_questions_to_exam(db: AsyncSession, exam_id: int, questions_data: list, is_course_exam: bool = False) -> None:
    """Add questions to an existing exam (admin)."""
    from app.modules.exams.models import EntranceExam, CourseExam, ExamQuestion, CourseExamQuestion, ExamOption, CourseExamOption
    
    exam_model = CourseExam if is_course_exam else EntranceExam
    question_model = CourseExamQuestion if is_course_exam else ExamQuestion
    option_model = CourseExamOption if is_course_exam else ExamOption

    exam = await db.get(exam_model, exam_id)
    if exam is None:
        raise HTTPException(status_code=404, detail="Exam not found")

    for q_data in questions_data:
        question = question_model(
            exam_id=exam_id,
            question_text=q_data["question_text"],
            question_type=q_data.get("question_type", "mcq"),
            marks=q_data.get("marks", 1.0),
            negative_marks=q_data.get("negative_marks", 0.0),
            category=q_data.get("category"),
            order=q_data.get("order", 0),
            explanation=q_data.get("explanation"),
        )
        db.add(question)
        await db.flush()

        for opt_data in q_data.get("options", []):
            option = option_model(
                question_id=question.id,
                option_text=opt_data["option_text"],
                is_correct=opt_data.get("is_correct", False),
                order=opt_data.get("order", 0),
            )
            db.add(option)

    await db.flush()
    logger.info("questions_added", exam_id=exam_id, count=len(questions_data))


async def parse_questions_from_file(file_bytes: bytes, filename: str) -> list:
    """Parse CSV or Excel file into a list of question dicts ready for add_questions_to_exam.

    Expected columns (case-insensitive, flexible naming):
        question | option_a | option_b | option_c | option_d | correct_answer | marks | negative_marks | category | explanation

    `correct_answer` should be one of: A, B, C, D (or a, b, c, d).
    Extra option columns (option_e, option_f ...) are supported automatically.
    """
    import csv
    import io

    rows: list[dict] = []

    lower_name = filename.lower()
    if lower_name.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        header = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {header[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(header)}
            rows.append(row_dict)
        wb.close()
    elif lower_name.endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({k.strip().lower().replace(" ", "_"): (v or "").strip() for k, v in row.items()})
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload .csv, .xlsx, or .xls")

    if not rows:
        raise HTTPException(status_code=400, detail="The uploaded file contains no data rows.")

    questions = []
    option_letters = "abcdefghij"

    for idx, row in enumerate(rows):
        q_text = row.get("question") or row.get("question_text") or ""
        if not q_text:
            continue  # skip blank rows

        correct_raw = (row.get("correct_answer") or row.get("answer") or "a").strip().lower()

        # Collect option columns dynamically
        options = []
        for letter in option_letters:
            col_name = f"option_{letter}"
            val = row.get(col_name, "").strip()
            if val:
                options.append({
                    "option_text": val,
                    "is_correct": correct_raw == letter,
                    "order": len(options),
                })

        marks = 1.0
        try:
            marks = float(row.get("marks", 1.0) or 1.0)
        except (ValueError, TypeError):
            pass

        neg_marks = 0.0
        try:
            neg_marks = float(row.get("negative_marks", 0.0) or 0.0)
        except (ValueError, TypeError):
            pass

        questions.append({
            "question_text": q_text,
            "question_type": row.get("question_type", "mcq") or "mcq",
            "marks": marks,
            "negative_marks": neg_marks,
            "category": row.get("category") or None,
            "explanation": row.get("explanation") or None,
            "order": idx,
            "options": options,
        })

    if not questions:
        raise HTTPException(status_code=400, detail="No valid questions could be parsed from the file.")

    logger.info("questions_parsed_from_file", filename=filename, count=len(questions))
    return questions


# ── Student: get available entrance exams ────────────────────────────
async def get_entrance_exams(db: AsyncSession) -> List[EntranceExam]:
    """List all active entrance exams."""
    result = await db.execute(
        select(EntranceExam)
        .where(EntranceExam.is_active == True)  # noqa: E712
        .order_by(EntranceExam.created_at.desc())
    )
    return list(result.scalars().all())


# ── Student: start exam ─────────────────────────────────────────────
async def start_exam(db: AsyncSession, user_id: int, exam_id: int) -> dict:
    """Create a new attempt after checking eligibility."""
    exam = await db.get(EntranceExam, exam_id)
    if exam is None or not exam.is_active:
        raise HTTPException(status_code=404, detail="Exam not found or inactive")

    # Check in-progress attempt
    result = await db.execute(
        select(ExamAttempt).where(
            ExamAttempt.user_id == user_id,
            ExamAttempt.exam_id == exam_id,
            ExamAttempt.is_submitted == False,  # noqa: E712
        )
    )
    existing = result.scalar_one_or_none()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="You already have an in-progress attempt for this exam.",
        )

    await _check_reattempt_allowed(db, user_id, exam_id)

    # Count questions
    q_count_result = await db.execute(
        select(func.count(ExamQuestion.id)).where(ExamQuestion.exam_id == exam_id)
    )
    total_questions = q_count_result.scalar() or 0

    attempt = ExamAttempt(exam_id=exam_id, user_id=user_id)
    db.add(attempt)
    await db.flush()
    await db.refresh(attempt)
    logger.info("exam_started", user_id=user_id, exam_id=exam_id, attempt_id=attempt.id)

    # Determine how many questions the student will see
    display_count = total_questions
    if exam.questions_per_attempt and exam.questions_per_attempt > 0:
        display_count = min(exam.questions_per_attempt, total_questions)

    return {
        "attempt_id": attempt.id,
        "exam_id": exam_id,
        "started_at": attempt.started_at,
        "duration_minutes": exam.duration_minutes,
        "total_questions": display_count,
    }


# ── Student: get questions for an attempt ────────────────────────────
async def get_exam_questions(db: AsyncSession, user_id: int, exam_id: int) -> List[ExamQuestion]:
    """Return questions with options (without correct-answer flags) for the student.
    
    If the exam has `questions_per_attempt` set, a random subset is returned.
    The random selection is consistent per attempt (seeded by attempt id).
    """
    import random as _random

    # Verify active attempt exists
    result = await db.execute(
        select(ExamAttempt).where(
            ExamAttempt.user_id == user_id,
            ExamAttempt.exam_id == exam_id,
            ExamAttempt.is_submitted == False,  # noqa: E712
        )
    )
    attempt = result.scalar_one_or_none()
    if attempt is None:
        raise HTTPException(status_code=400, detail="No active attempt found. Start the exam first.")

    q_result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(ExamQuestion.exam_id == exam_id)
        .order_by(ExamQuestion.order)
    )
    all_questions = list(q_result.scalars().all())

    # Check if we need to randomize / subset
    exam = await db.get(EntranceExam, exam_id)
    if exam and exam.questions_per_attempt and exam.questions_per_attempt > 0 and len(all_questions) > exam.questions_per_attempt:
        rng = _random.Random(attempt.id)  # deterministic per attempt so refreshing gives same set
        all_questions = rng.sample(all_questions, exam.questions_per_attempt)

    return all_questions


# ── Student: save individual answer ─────────────────────────────────
async def save_answer(
    db: AsyncSession, user_id: int, attempt_id: int, question_id: int, selected_option_id: Optional[int], descriptive_text: Optional[str] = None
) -> ExamAnswer:
    """Save or update an answer for a question within an attempt."""
    # Verify attempt belongs to user and is not submitted
    attempt = await db.get(ExamAttempt, attempt_id)
    if attempt is None or attempt.user_id != user_id or attempt.is_submitted:
        raise HTTPException(status_code=400, detail="Invalid or submitted attempt")

    # Check if answer already exists
    result = await db.execute(
        select(ExamAnswer).where(
            ExamAnswer.attempt_id == attempt_id,
            ExamAnswer.question_id == question_id,
        )
    )
    existing = result.scalar_one_or_none()

    if existing:
        existing.selected_option_id = selected_option_id
        existing.descriptive_text = descriptive_text
        existing.answered_at = datetime.now(timezone.utc)
        await db.flush()
        return existing

    answer = ExamAnswer(
        attempt_id=attempt_id,
        question_id=question_id,
        selected_option_id=selected_option_id,
        descriptive_text=descriptive_text,
    )
    db.add(answer)
    await db.flush()
    return answer


# ── Student: submit exam and evaluate ────────────────────────────────
async def submit_exam(db: AsyncSession, user_id: int, attempt_id: int, answers: list) -> ExamResult:
    """Submit the exam, evaluate answers, and generate result."""
    attempt = await db.get(ExamAttempt, attempt_id)
    if attempt is None or attempt.user_id != user_id:
        raise HTTPException(status_code=400, detail="Invalid attempt")
    if attempt.is_submitted:
        raise HTTPException(status_code=409, detail="Exam already submitted")

    # Save any remaining answers
    for ans in answers:
        await save_answer(db, user_id, attempt_id, ans["question_id"], ans.get("selected_option_id"), ans.get("descriptive_text"))

    # Get all questions for this exam
    q_result = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(ExamQuestion.exam_id == attempt.exam_id)
    )
    questions = list(q_result.scalars().all())

    # Get all student answers for this attempt
    a_result = await db.execute(
        select(ExamAnswer).where(ExamAnswer.attempt_id == attempt_id)
    )
    student_answers = {a.question_id: a for a in a_result.scalars().all()}

    # Evaluate
    total_marks = 0.0
    obtained_marks = 0.0
    correct_count = 0

    needs_manual_evaluation = False

    for question in questions:
        q_marks = question.marks if question.marks is not None else 1.0
        q_neg = question.negative_marks if question.negative_marks is not None else 0.0
        total_marks += q_marks
        student_answer = student_answers.get(question.id)
        
        if question.question_type == "descriptive":
            needs_manual_evaluation = True
            continue

        if student_answer and student_answer.selected_option_id:
            # Find the correct option
            correct_option = next((o for o in question.options if o.is_correct), None)
            if correct_option and student_answer.selected_option_id == correct_option.id:
                student_answer.is_correct = True
                obtained_marks += q_marks
                correct_count += 1
            else:
                student_answer.is_correct = False
                obtained_marks -= q_neg

    percentage = (obtained_marks / total_marks * 100) if total_marks > 0 else 0.0

    # Get passing score
    exam = await db.get(EntranceExam, attempt.exam_id)
    passed = percentage >= exam.passing_score

    # Mark attempt as submitted
    now = datetime.now(timezone.utc)
    attempt.is_submitted = True
    attempt.needs_manual_evaluation = needs_manual_evaluation
    attempt.submitted_at = now
    if attempt.started_at:
        attempt.time_spent_seconds = int((now - attempt.started_at).total_seconds())

    # Create result
    exam_result = ExamResult(
        attempt_id=attempt_id,
        user_id=user_id,
        exam_id=attempt.exam_id,
        total_questions=len(questions),
        correct_answers=correct_count,
        total_marks=total_marks,
        obtained_marks=obtained_marks,
        percentage=round(percentage, 2),
        passed=passed,
    )
    db.add(exam_result)
    await db.flush()
    await db.refresh(exam_result)

    logger.info(
        "exam_submitted",
        user_id=user_id,
        attempt_id=attempt_id,
        percentage=percentage,
        passed=passed,
    )
    return exam_result


# ── Student: get result ─────────────────────────────────────────────
async def get_exam_result(db: AsyncSession, user_id: int, exam_id: int) -> Optional[ExamResult]:
    """Get the most recent exam result for a user."""
    result = await db.execute(
        select(ExamResult)
        .where(ExamResult.user_id == user_id, ExamResult.exam_id == exam_id)
        .order_by(ExamResult.evaluated_at.desc())
        .limit(1)
    )
    exam_result = result.scalar_one_or_none()
    if exam_result is None:
        raise HTTPException(status_code=404, detail="No exam result found")
    return exam_result


# ── Phase 2: Course & Monthly Exams ───────────────────────────────────

from app.modules.exams.models import (
    CourseExam, MonthlyExam, ExamPayment, CourseExamAttempt, ExamViolation,
    CourseExamQuestion, CourseExamOption, CourseExamAnswer, CourseExamResult,
)


async def submit_course_exam(db: AsyncSession, user_id: int, attempt_id: int, answers: list) -> CourseExamResult:
    """Submit a course exam, evaluate answers, and generate result."""
    attempt = await db.get(CourseExamAttempt, attempt_id)
    if attempt is None or attempt.user_id != user_id:
        raise HTTPException(status_code=400, detail="Invalid attempt")
    if attempt.is_submitted:
        raise HTTPException(status_code=409, detail="Exam already submitted")

    # Save answers
    for ans in answers:
        q_id = ans["question_id"]
        opt_id = ans.get("selected_option_id")
        desc = ans.get("descriptive_text")
        if opt_id or desc:
            db.add(CourseExamAnswer(
                attempt_id=attempt_id, question_id=q_id,
                selected_option_id=opt_id, descriptive_text=desc
            ))
    await db.flush()

    # Get all questions for this exam
    q_result = await db.execute(
        select(CourseExamQuestion)
        .options(selectinload(CourseExamQuestion.options))
        .where(CourseExamQuestion.exam_id == attempt.exam_id)
    )
    questions = list(q_result.scalars().all())

    # Get all student answers for this attempt
    a_result = await db.execute(
        select(CourseExamAnswer).where(CourseExamAnswer.attempt_id == attempt_id)
    )
    student_answers = {a.question_id: a for a in a_result.scalars().all()}

    # Evaluate
    total_marks = 0.0
    obtained_marks = 0.0
    correct_count = 0

    for question in questions:
        q_marks = question.marks if question.marks is not None else 1.0
        q_neg = question.negative_marks if question.negative_marks is not None else 0.0
        total_marks += q_marks
        student_answer = student_answers.get(question.id)

        if not student_answer or not student_answer.selected_option_id:
            continue

        correct_option = next((o for o in question.options if o.is_correct), None)
        if correct_option and student_answer.selected_option_id == correct_option.id:
            student_answer.is_correct = True
            obtained_marks += q_marks
            correct_count += 1
        else:
            student_answer.is_correct = False
            obtained_marks -= q_neg

    percentage = (obtained_marks / total_marks * 100) if total_marks > 0 else 0.0

    exam = await db.get(CourseExam, attempt.exam_id)
    passed = percentage >= (exam.passing_score if exam else 60)

    # Mark attempt as submitted
    now = datetime.now(timezone.utc)
    attempt.is_submitted = True
    attempt.submitted_at = now
    if attempt.started_at:
        attempt.time_spent_seconds = int((now - attempt.started_at).total_seconds())

    # Create result
    exam_result = CourseExamResult(
        attempt_id=attempt_id,
        user_id=user_id,
        exam_id=attempt.exam_id,
        total_questions=len(questions),
        correct_answers=correct_count,
        total_marks=total_marks,
        obtained_marks=obtained_marks,
        percentage=round(percentage, 2),
        passed=passed,
    )
    db.add(exam_result)
    await db.flush()
    await db.refresh(exam_result)

    logger.info(
        "course_exam_submitted",
        user_id=user_id,
        attempt_id=attempt_id,
        percentage=percentage,
        passed=passed,
    )
    return exam_result

async def get_course_exam_result(db: AsyncSession, user_id: int, exam_id: int) -> Optional[CourseExamResult]:
    """Get the most recent course exam result for a user."""
    result = await db.execute(
        select(CourseExamResult)
        .where(CourseExamResult.user_id == user_id, CourseExamResult.exam_id == exam_id)
        .order_by(CourseExamResult.evaluated_at.desc())
        .limit(1)
    )
    exam_result = result.scalar_one_or_none()
    if exam_result is None:
        raise HTTPException(status_code=404, detail="No exam result found")
    return exam_result


async def get_admin_exam_questions(db: AsyncSession, exam_id: int, is_course: bool):
    """Fetch all questions for an exam (for teacher/admin to view in Question Builder)."""
    if is_course:
        result = await db.execute(
            select(CourseExamQuestion)
            .options(selectinload(CourseExamQuestion.options))
            .where(CourseExamQuestion.exam_id == exam_id)
            .order_by(CourseExamQuestion.order)
        )
    else:
        result = await db.execute(
            select(ExamQuestion)
            .options(selectinload(ExamQuestion.options))
            .where(ExamQuestion.exam_id == exam_id)
            .order_by(ExamQuestion.order)
        )
    return list(result.scalars().all())

async def get_monthly_exams(db: AsyncSession, user_id: int) -> List[MonthlyExam]:
    """Fetch all monthly exams."""
    result = await db.execute(
        select(MonthlyExam).options(selectinload(MonthlyExam.exam))
    )
    return list(result.scalars().all())

async def process_exam_payment(db: AsyncSession, user_id: int, exam_id: int, amount: float) -> ExamPayment:
    """Mock payment for reattempt."""
    payment = ExamPayment(user_id=user_id, exam_id=exam_id, amount=amount, status="paid")
    db.add(payment)
    await db.flush()
    return payment

async def verify_course_exam_attempt_allowed(db: AsyncSession, user_id: int, exam_id: int) -> None:
    """Check if the user can start a new course exam attempt."""
    from app.modules.exams.models import CourseExamResult

    # Check for in-progress attempts — auto-close stale ones
    in_progress_res = await db.execute(
        select(CourseExamAttempt).where(
            CourseExamAttempt.user_id == user_id,
            CourseExamAttempt.exam_id == exam_id,
            CourseExamAttempt.is_submitted == False,  # noqa: E712
        ).order_by(CourseExamAttempt.started_at.desc())
    )
    in_progress_attempts = list(in_progress_res.scalars().all())
    
    if in_progress_attempts:
        # Get exam duration to determine staleness
        exam = await db.get(CourseExam, exam_id)
        cutoff = datetime.now(timezone.utc) - timedelta(minutes=(exam.duration_minutes * 2 if exam else 120))
        
        active_attempt = None
        for attempt in in_progress_attempts:
            if attempt.started_at and attempt.started_at < cutoff:
                # Auto-close stale attempt
                attempt.is_submitted = True
                attempt.submitted_at = datetime.now(timezone.utc)
            else:
                active_attempt = attempt
        await db.flush()
        
        if active_attempt:
            raise HTTPException(status_code=409, detail="You already have an in-progress attempt for this exam.")

    # Check submitted attempts — eagerly load result
    result = await db.execute(
        select(CourseExamAttempt)
        .options(selectinload(CourseExamAttempt.result))
        .where(
            CourseExamAttempt.user_id == user_id,
            CourseExamAttempt.exam_id == exam_id,
            CourseExamAttempt.is_submitted == True,  # noqa: E712
        )
        .order_by(CourseExamAttempt.submitted_at.desc())
        .limit(1)
    )
    last_attempt = result.scalar_one_or_none()
    if not last_attempt:
        return  # First attempt — always allowed

    # Already passed — no need to retake
    if last_attempt.result and last_attempt.result.passed:
        raise HTTPException(status_code=409, detail="You have already passed this exam.")

    # Failed previously — check max attempts on the exam
    exam = await db.get(CourseExam, exam_id)
    if exam and exam.max_attempts > 0:
        # Count total submitted attempts
        count_res = await db.execute(
            select(func.count(CourseExamAttempt.id)).where(
                CourseExamAttempt.user_id == user_id,
                CourseExamAttempt.exam_id == exam_id,
                CourseExamAttempt.is_submitted == True,  # noqa: E712
            )
        )
        total = count_res.scalar() or 0
        if total >= exam.max_attempts:
            # Check if payment was made for reattempt
            pay_result = await db.execute(
                select(ExamPayment).where(
                    ExamPayment.user_id == user_id,
                    ExamPayment.exam_id == exam_id,
                    ExamPayment.status == "paid"
                )
            )
            if not pay_result.scalars().first():
                raise HTTPException(
                    status_code=402,
                    detail=f"You have used all {exam.max_attempts} attempts. Payment required to reattempt."
                )


async def start_course_exam(db: AsyncSession, user_id: int, exam_id: int, device_id: str) -> dict:
    exam = await db.get(CourseExam, exam_id)
    if not exam:
        raise HTTPException(status_code=404, detail="Exam not found")
    if not exam.is_active:
        raise HTTPException(status_code=400, detail="This exam is not currently active.")
        
    await verify_course_exam_attempt_allowed(db, user_id, exam_id)
    
    attempt = CourseExamAttempt(user_id=user_id, exam_id=exam_id, device_id=device_id)
    db.add(attempt)
    await db.flush()
    
    return {
        "attempt_id": attempt.id,
        "exam_id": exam_id,
        "started_at": attempt.started_at,
        "duration_minutes": exam.duration_minutes,
        "device_id": attempt.device_id
    }

async def get_course_exam_questions(db: AsyncSession, user_id: int, exam_id: int):
    """Return course exam questions for the student's active attempt.
    
    If the exam has `questions_per_attempt` set, a random subset is returned.
    The selection is deterministic per attempt (seeded by attempt id).
    """
    import random as _random
    from app.modules.exams.models import CourseExamQuestion, CourseExamOption

    # Verify active course exam attempt exists
    result = await db.execute(
        select(CourseExamAttempt).where(
            CourseExamAttempt.user_id == user_id,
            CourseExamAttempt.exam_id == exam_id,
            CourseExamAttempt.is_submitted == False,  # noqa: E712
        )
    )
    attempt = result.scalar_one_or_none()
    if attempt is None:
        raise HTTPException(status_code=400, detail="No active attempt found. Start the exam first.")

    q_result = await db.execute(
        select(CourseExamQuestion)
        .options(selectinload(CourseExamQuestion.options))
        .where(CourseExamQuestion.exam_id == exam_id)
        .order_by(CourseExamQuestion.order)
    )
    all_questions = list(q_result.scalars().all())

    # Check if we need to randomize / subset
    exam = await db.get(CourseExam, exam_id)
    if exam and exam.questions_per_attempt and exam.questions_per_attempt > 0 and len(all_questions) > exam.questions_per_attempt:
        rng = _random.Random(attempt.id)
        all_questions = rng.sample(all_questions, exam.questions_per_attempt)

    return all_questions

async def log_exam_violation(db: AsyncSession, user_id: int, attempt_id: int, violation_type: str) -> None:
    violation = ExamViolation(attempt_id=attempt_id, violation_type=violation_type)
    db.add(violation)
    await db.flush()

async def close_exam_session(db: AsyncSession, user_id: int, attempt_id: int) -> None:
    attempt = await db.get(CourseExamAttempt, attempt_id)
    if attempt and not attempt.is_submitted and attempt.user_id == user_id:
        attempt.is_submitted = True
        attempt.submitted_at = datetime.now(timezone.utc)
        await db.flush()


# ── Skill-Based Result Analysis ──────────────────────────────────────

from app.modules.exams.models import CategoryScore

async def get_skill_analysis(db: AsyncSession, user_id: int) -> dict:
    """Analyse user's category scores and return strong/weak areas."""
    result = await db.execute(
        select(CategoryScore).where(CategoryScore.user_id == user_id)
    )
    scores = list(result.scalars().all())

    if not scores:
        return {"strong_areas": [], "weak_areas": [], "suggestions": ["Complete exams to see your skill analysis."]}

    scored = []
    for s in scores:
        pct = (s.score / s.max_score * 100) if s.max_score > 0 else 0
        scored.append({
            "category": s.category,
            "score": s.score,
            "max_score": s.max_score,
            "percentage": round(pct, 1),
        })

    scored.sort(key=lambda x: x["percentage"], reverse=True)

    strong = [s for s in scored if s["percentage"] >= 70]
    weak = [s for s in scored if s["percentage"] < 70]

    suggestions = []
    for w in weak:
        suggestions.append(f"Improve your {w['category']} skills — currently at {w['percentage']}%.")
    if not weak:
        suggestions.append("Great job! You're performing well across all categories.")

    return {"strong_areas": strong, "weak_areas": weak, "suggestions": suggestions}

async def get_course_exam_attempt_review(db: AsyncSession, user_id: int, attempt_id: int) -> dict:
    """Get detailed review for a specific course exam attempt."""
    from app.modules.exams.models import CourseExamAttempt, CourseExamAnswer, CourseExamQuestion, CourseExamResult, ExamViolation
    
    # 1. Fetch attempt with result and violations
    res = await db.execute(
        select(CourseExamAttempt)
        .options(
            selectinload(CourseExamAttempt.result),
            selectinload(CourseExamAttempt.violations),
            selectinload(CourseExamAttempt.exam)
        )
        .where(CourseExamAttempt.id == attempt_id, CourseExamAttempt.user_id == user_id)
    )
    attempt = res.scalar_one_or_none()
    if not attempt:
        raise HTTPException(404, "Attempt not found")
    
    # 2. Fetch questions and student answers
    q_res = await db.execute(
        select(CourseExamQuestion)
        .options(selectinload(CourseExamQuestion.options))
        .where(CourseExamQuestion.exam_id == attempt.exam_id)
        .order_by(CourseExamQuestion.order)
    )
    questions = list(q_res.scalars().all())
    
    a_res = await db.execute(
        select(CourseExamAnswer).where(CourseExamAnswer.attempt_id == attempt_id)
    )
    answers = {a.question_id: a for a in a_res.scalars().all()}
    
    # 3. Format response
    review_questions = []
    for q in questions:
        ans = answers.get(q.id)
        review_questions.append({
            "id": q.id,
            "question_text": q.question_text,
            "question_type": q.question_type,
            "marks": q.marks,
            "negative_marks": q.negative_marks,
            "explanation": q.explanation,
            "options": [
                {"id": o.id, "option_text": o.option_text, "is_correct": o.is_correct}
                for o in q.options
            ],
            "selected_option_id": ans.selected_option_id if ans else None,
            "is_correct": ans.is_correct if ans else None
        })
        
    violations = [v.violation_type for v in attempt.violations]
    
    return {
        "attempt_id": attempt.id,
        "exam_title": attempt.exam.title,
        "started_at": attempt.started_at,
        "submitted_at": attempt.submitted_at,
        "total_questions": attempt.result.total_questions if attempt.result else 0,
        "correct_answers": attempt.result.correct_answers if attempt.result else 0,
        "total_marks": attempt.result.total_marks if attempt.result else 0.0,
        "obtained_marks": attempt.result.obtained_marks if attempt.result else 0.0,
        "percentage": attempt.result.percentage if attempt.result else 0.0,
        "passed": attempt.result.passed if attempt.result else False,
        "violations": violations,
        "questions": review_questions
    }

async def get_entrance_exam_attempt_review(db: AsyncSession, user_id: int, attempt_id: int) -> dict:
    """Get detailed review for a specific entrance exam attempt."""
    from app.modules.exams.models import ExamAttempt, ExamAnswer, ExamQuestion, ExamResult
    
    # 1. Fetch attempt with result
    res = await db.execute(
        select(ExamAttempt)
        .options(
            selectinload(ExamAttempt.result),
            selectinload(ExamAttempt.exam)
        )
        .where(ExamAttempt.id == attempt_id, ExamAttempt.user_id == user_id)
    )
    attempt = res.scalar_one_or_none()
    if not attempt:
        raise HTTPException(404, "Attempt not found")
    
    # 2. Fetch questions and student answers
    q_res = await db.execute(
        select(ExamQuestion)
        .options(selectinload(ExamQuestion.options))
        .where(ExamQuestion.exam_id == attempt.exam_id)
        .order_by(ExamQuestion.order)
    )
    questions = list(q_res.scalars().all())
    
    a_res = await db.execute(
        select(ExamAnswer).where(ExamAnswer.attempt_id == attempt_id)
    )
    answers = {a.question_id: a for a in a_res.scalars().all()}
    
    # 3. Format response
    review_questions = []
    for q in questions:
        ans = answers.get(q.id)
        review_questions.append({
            "id": q.id,
            "question_text": q.question_text,
            "question_type": q.question_type,
            "marks": q.marks,
            "negative_marks": q.negative_marks,
            "explanation": q.explanation,
            "options": [
                {"id": o.id, "option_text": o.option_text, "is_correct": o.is_correct}
                for o in q.options
            ],
            "selected_option_id": ans.selected_option_id if ans else None,
            "is_correct": ans.is_correct if ans else None
        })
        
    return {
        "attempt_id": attempt.id,
        "exam_title": attempt.exam.title,
        "started_at": attempt.started_at,
        "submitted_at": attempt.submitted_at,
        "total_questions": attempt.result.total_questions if attempt.result else 0,
        "correct_answers": attempt.result.correct_answers if attempt.result else 0,
        "total_marks": attempt.result.total_marks if attempt.result else 0.0,
        "obtained_marks": attempt.result.obtained_marks if attempt.result else 0.0,
        "percentage": attempt.result.percentage if attempt.result else 0.0,
        "passed": attempt.result.passed if attempt.result else False,
        "violations": [],
        "questions": review_questions
    }
